# -*- coding: utf-8 -*-
"""
This script is for python 3.x

"""
# file path issue
# bold cell
# write to worksheet: a general one and an agency one
#some of the bold cell are in function (zero_one_accept_missing, check_baseline)

import os
import os.path
import sys
import openpyxl
import shutil
import _pickle as pickle
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.comments import Comment
from openpyxl.styles.borders import Side
from copy import copy 
import pandas as pd
from datetime import datetime

def mk_int(s):
    s = str(s).strip()
    return int(float(s)) if s else None
         
def employment_date_check(data,row_number,column_number):
    if (data.cell(row = row_number, column = column_number).value == 3) or (data.cell(row = row_number, column = column_number).value == "3"):
        sd_value = data.cell(row = row_number, column = column_number+1).value
        if (sd_value is None) or (sd_value == " ") or (sd_value == ""): return 0
        else: # bold both start date and employment status cell
            data.cell(row = row_number, column = column_number).fill = RED
            data.cell(row = row_number, column = column_number+1).fill = RED
            data.cell(row = row_number, column = column_number+1).comment = unemployment_but_value
            data.cell(row = row_number, column = last_column+issues+1).value = "{}:The client is unemployed but has value here, please check typing/coding".format(data.cell(row = 1, column = column_number+1).value)            
            return 1
    elif (data.cell(row = row_number, column = column_number).value == 2) or \
    (data.cell(row = row_number, column = column_number).value == "2") or \
    (data.cell(row = row_number, column = column_number).value == 1) or \
    (data.cell(row = row_number, column = column_number).value == "1"):
        sd_value = data.cell(row = row_number, column = column_number+1).value
        if (sd_value is None) or (sd_value == " ") or (sd_value == ""):
            data.cell(row = row_number, column = column_number).fill = RED
            data.cell(row = row_number, column = column_number+1).fill = RED
            data.cell(row = row_number, column = column_number+1).comment = employment_but_no_value
            data.cell(row = row_number, column = last_column+issues+1).value = "{}:The client is employed but has no value here, please check typing/coding".format(data.cell(row = 1, column = column_number+1).value)
            # bold both start date and employment status cell
            return 1
        else: return 0
    else: return 0
        
def employment_wage_check(data,row_number,column_number):
    if (data.cell(row = row_number, column = column_number).value == 3) or (data.cell(row = row_number, column = column_number).value == "3"):
        wg1 = data.cell(row = row_number, column = column_number+9).value
        wg2 = data.cell(row = row_number, column = column_number+11).value
        if ((wg1 is None) or (wg1 == " ") or (wg1 ==" ") or (wg1 == 0) or (wg1 == "0") or (wg1 =="")) and \
        ((wg2 is None) or (wg2 == " ") or (wg2 ==" ") or (wg2 == 0) or (wg2 == "0") or (wg2 =="")):
            return 0
        else: #bold both wages and employment status
            data.cell(row = row_number, column = column_number).fill = RED
            data.cell(row = row_number, column = column_number+9).fill = RED
            data.cell(row = row_number, column = column_number+11).fill = RED
            data.cell(row = row_number, column = column_number+9).comment = unemployment_but_value
            data.cell(row = row_number, column = column_number+11).comment = unemployment_but_value
            data.cell(row = row_number, column = last_column+issues+1).value = "{},{}:The client is unemployed but has value in at least one of the wage columns, please check typing/coding".format(data.cell(row = 1, column = column_number+9).value,data.cell(row = 1, column = column_number+11).value)
            return 1
    elif (data.cell(row = row_number, column = column_number).value == 2) or \
    (data.cell(row = row_number, column = column_number).value == "2") or \
    (data.cell(row = row_number, column = column_number).value == 1) or \
    (data.cell(row = row_number, column = column_number).value == "1"):
        wg1 = data.cell(row = row_number, column = column_number+9).value
        wg2 = data.cell(row = row_number, column = column_number+11).value        
        if ((wg1 is None) or (wg1 == " ") or (wg1 ==" ") or (wg1 == 0) or (wg1 == "0") or (wg1 =="")) and \
        ((wg2 is None) or (wg2 == " ") or (wg2 ==" ") or (wg2 == 0) or (wg2 == "0") or (wg2 =="")):
            #bold both wages and employment status
            data.cell(row = row_number, column = column_number).fill = RED
            data.cell(row = row_number, column = column_number+9).fill = RED
            data.cell(row = row_number, column = column_number+11).fill = RED
            data.cell(row = row_number, column = column_number+9).comment = employment_but_no_value
            data.cell(row = row_number, column = column_number+11).comment = employment_but_no_value
            data.cell(row = row_number, column = last_column+issues+1).value = "{},{}:The client is employed but has no value in neither of the wage columns, please check typing/coding".format(data.cell(row = 1, column = column_number+9).value,data.cell(row = 1, column = column_number+11).value)
            return 1
        else: return 0
    else: return 0
              
def zero_one_accept_missing(data, rownumber, columnnumber):
    if data.cell(row = rownumber, column = columnnumber).value is not None:
        if (type(data.cell(row = rownumber, column = columnnumber).value) == int) or (type(data.cell(row = rownumber, column = columnnumber).value) == float):
            a = mk_int(data.cell(row = rownumber, column = columnnumber).value)
            if a not in [0,1]:
                #bold the cell
                data.cell(row = rownumber, column = columnnumber).fill = RED
                data.cell(row = rownumber, column = columnnumber).comment = number_out_of_range
                data.cell(row = rownumber, column = last_column+issues+1).value = "{}:The number is out of range".format(data.cell(row = 1, column = columnnumber).value)
                return 1
            else: return 0
        else: 
            if data.cell(row = rownumber, column = columnnumber).value.strip() != "":
                try:
                    a = mk_int(data.cell(row = rownumber, column = columnnumber).value)
                    if a not in [0,1]:
                        #bold the cell
                        data.cell(row = rownumber, column = columnnumber).fill = RED
                        data.cell(row = rownumber, column = columnnumber).comment = string_out_of_range
                        data.cell(row = rownumber, column = last_column+issues+1).value = "{}:The data type of this cell is string, and the value is out of range".format(data.cell(row = 1, column = columnnumber).value)
                        return 1
                    else: return 0
                except: 
                    #bold the cell, the value is not number
                    data.cell(row = rownumber, column = columnnumber).fill = RED
                    data.cell(row = rownumber, column = columnnumber).comment = value_wrong
                    data.cell(row = rownumber, column = last_column+issues+1).value = "{}:The value is not acceptable here, please check typing/coding".format(data.cell(row = 1, column = columnnumber).value)
                    return 1               
            else: return 0                
    else: return 0
    
    
def check_int_float(data, rownumber, columnnumber):
    if data.cell(row = rownumber, column = columnnumber).value is not None:
        if (type(data.cell(row = rownumber, column = columnnumber).value) == int) or (type(data.cell(row = rownumber, column = columnnumber).value) == float):
            return 0
        else:
            if data.cell(row = rownumber, column = columnnumber).value.strip()=="":
                return 0
            else:
                try:
                    #data type issue: int stored as str
                    a = mk_int(data.cell(row = rownumber, column = columnnumber).value)
                    dn = columnnumber - 1
                    data_type_issue[column_names[dn]] = 1
                    return 0
                except:
                    #bord the cell, data issue
                    data.cell(row = rownumber, column = columnnumber).fill = RED
                    data.cell(row = rownumber, column = columnnumber).comment = not_a_number
                    data.cell(row = rownumber, column = last_column+issues+1).value = "{}:The value of this cell is not a numeric number".format(data.cell(row = 1, column = columnnumber).value)
                    dn = columnnumber - 1
                    data_type_issue[column_names[dn]] = 1
                    return 1
    else: return 0 
            



def check_baseline(data, row_number):
    if str(data.cell(row = row_number, column = 2).value) == "1":
        if str(data.cell(row = row_number, column =1).value).replace(",","") in clientid:
            # bold
            data.cell(row = row_number, column = 2).fill = RED
            data.cell(row = row_number, column = 2).comment = baseline_one_wrong
            data.cell(row = row_number, column = last_column+issues+1).value = "{}:This client id exists in baseline database.".format(data.cell(row = 1, column = 2).value)
            return 1
        else: return 0
    elif str(data.cell(row = row_number, column = 2).value) == "0":
        if str(data.cell(row = row_number, column =1).value).replace(",","") not in clientid:
            # bold
            data.cell(row = row_number, column = 2).fill = RED
            data.cell(row = row_number, column = 2).comment = baseline_zero_wrong
            data.cell(row = row_number, column = last_column+issues+1).value = "{}:This client id does not exist in baseline database.".format(data.cell(row = 1, column = 2).value)
            return 1
        else: return 0
    else: 
        return 0
        

def Get_Agency_Name(file):
    for i in AgencyName:
        if i.lower() in file.lower():
            Agency = i
        else: continue
    return Agency


def Get_Sheet_Name(filepath,agencyname):
    sheet_name = ""
    if agencyname in ["CCD","CitySquare"]:
        xl = pd.ExcelFile(filepath)
        sheet_name = xl.sheet_names[0]
    elif agencyname in ["HCC","JFS","Metrocrest","WCTC"]:
        sheet_name = "Data Submission Basic Template"
    elif agencyname == "Interfaith":
        sheet_name = "Data Submission Detail Template"
    elif agencyname =="IRC":
        xl = pd.ExcelFile(filepath)
        sheet_name = xl.sheet_names[1]
    else: 
        sheet_name = ""
        print("Undetected file: ",filepath)
    return sheet_name
    
def read_excel_file(filepath, sheetname):
        wb = load_workbook(filepath,data_only=True)
        ws = wb[sheetname]
        return wb, ws

def alarming_str(table,row_number,column_number):
    if table.cell(row = row_number,column = column_number).value is not None:
        if (type(table.cell(row = row_number, column = column_number).value) == int) or (type(table.cell(row = row_number, column = column_number).value) == float):
            pass
        else:
            if table.cell(row = row_number, column = column_number).value.strip() !="":
                dn = column_number - 1
                data_type_issue[column_names[dn]] = 1
            else: pass
    else: pass


#yellow
YELLOW = PatternFill(start_color="FFEEE78C", end_color="FFEEE78C", fill_type = "solid")
#red
ORANGE = PatternFill(start_color='FFEE7451',end_color='FFEE7451',fill_type='solid')
RED = YELLOW
baseline_zero_wrong=Comment("No baseline record found","MS")
baseline_one_wrong=Comment("Record exists in baseline database","MS")
number_out_of_range=Comment("The number is out of range","MS")
string_out_of_range=Comment("The data type is string, and value is out of range","MS")
value_wrong=Comment("Contains value issue, unacceptable value","MS")
not_a_number=Comment("Numeric numbers not detected","MS")
missing_value=Comment("Missing value", "MS")
unemployment_but_value=Comment("Client unemployed, but contains value other than None or 0","MS")
employment_but_no_value=Comment("Client employed, but miss value here","MS")
AgencyName = ["CCD","CitySquare","HCC","Interfaith","IRC",
              "JFS","Metrocrest","WCTC"]

try:
    os.chdir('C:\\Users\\Qiao Zhang\\Desktop\\Communities Foundation of Texas\\2018Q4')
    print("Working Directory: " + os.getcwd())
except:
    print("failure in changing working directory")
    sys.exit()
IN_FILE_PATH = os.path.abspath("Data Submission") # = os.path.join(os.getcwd(),"Sub")
timenow = str(datetime.now())[2:10]
CWD = os.getcwd()
try:
    os.mkdir(f'{CWD}/DataCheckResult{timenow}')
    result_folder = f'{CWD}/DataCheckResult{timenow}'
    print("make directory successfully")
except:
    try:
        DataLog.close()
    except: pass
    shutil.rmtree(f'{CWD}/DataCheckResult{timenow}', ignore_errors=True)
    os.mkdir(f'{CWD}/DataCheckResult{timenow}')
    result_folder = f'{CWD}/DataCheckResult{timenow}'
    print("remake directory "+f'{CWD}/DataCheckResult{timenow}')
    #sys.exit()


#os.path.abspath("clientid_pickle")
h = open(f'{CWD}/client_id_pickle','rb')
pickling = pickle.load(h)
clientid = pickling['clientid_baseline']
clientid = clientid.astype(str)
h.close()

DataLog = open(f'{CWD}/DataCheckResult{timenow}/Data_Check_Log.txt','w')

for dirpath, dirs, files in os.walk(os.path.normpath(IN_FILE_PATH),topdown = True):
    for file in files:
        #print(dirpath)
        DataLog.write("-"*80+"\n")
        DataLog.write(file+"\n")
        FILE_DETAIL_PATH = os.path.join(dirpath, file)
        #print(FILE_DETAIL_PATH)
        Agency_Name = Get_Agency_Name(file)
        My_Sheet_Name = Get_Sheet_Name(FILE_DETAIL_PATH,Agency_Name)
        if not My_Sheet_Name:
            # add to log
            DataLog.write("Agency name is not detected in {}, file skipped".format(file) +"\n")
            continue
        wb, excel = read_excel_file(FILE_DETAIL_PATH, My_Sheet_Name)
        print(file)
        max_col = excel.max_column
        max_row = excel.max_row
        DataLog.write("{} columns and {} rows are detected by machine".format(max_col,max_row)+"\n")
        # code detect xxx columns, code detect xxx rows
        while excel.cell(row = 1, column = 1).value is None:
            excel.delete_cols(1)
        while excel.cell(row = 1, column = 1).value.strip().replace(" ","").lower() != 'clientid':
            excel.delete_cols(1)
        if My_Sheet_Name =="Data Submission Detail Template": excel.delete_rows(2,3)
        
        for i in range(1, max_col+1):
            if excel.cell(row = 1, column =i).value == 'FinCap_LateFee':
                last_column = i
            else: pass
        #Get column names
        column_names = []
        for i in range(1, max_col +1):
            cell_obj = excel.cell(row = 1, column = i)
            column_names.append(cell_obj.value)
        #Get row numbers
        real_max_row = 0
        for i in range(1,max_row+1):
            if excel.cell(row = i, column = 1).value is not None:
                real_max_row = real_max_row + 1
            else:
                break      
        DataLog.write("Numbers of real row (not null): {}".format(real_max_row) +"\n")
        DataLog.write("Column Names: {}".format(column_names) +"\n")
        if last_column == 39:
            total_discrepancy_column = 0
            sheet_issues = 0
            wrong_info = 0
            data_type_issue = {}
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.worksheets[0]
            for i in range(2, real_max_row+1):
                issues = 0
                alarming_str(excel,i,1)
                alarming_str(excel,i,2)
                #need a bold here
                issues = issues + check_baseline(excel,i)
                alarming_str(excel,i,13)
                # Check Employment Service Access
                issues = issues +zero_one_accept_missing(excel,i,13)
                #Check Employment Status
                alarming_str(excel,i,14)
                if excel.cell(row = i, column = 14).value is not None:
                    a = mk_int(excel.cell(row = i, column = 14).value)
                    if a:
                        if a not in [1,2,3]:
                        #bold the cell
                            excel.cell(row = i, column = 14).fill =RED
                            excel.cell(row = i, column = 14).comment =number_out_of_range
                            excel.cell(row = i, column = last_column+issues+1).value = "{}:The value of number is out of range".format(excel.cell(row = 1, column = 14).value)
                            issues = issues + 1
                        else: pass
                    else: 
                        #Flag it as issue (blank cell), need bold
                        excel.cell(row = i, column = 14).fill = RED
                        excel.cell(row = i, column = 14).comment = missing_value
                        excel.cell(row = i, column = last_column+issues+1).value = "{}:Missing value".format(excel.cell(row = 1, column = 14).value)
                        issues = issues + 1
                else:
                    #Flag it as issue (blank cell), need bold
                    excel.cell(row = i, column = 14).fill =RED
                    excel.cell(row = i, column = 14).comment = missing_value
                    excel.cell(row = i, column = last_column+issues+1).value = "{}:Missing value".format(excel.cell(row = 1, column = 14).value)
                    issues = issues +1
                #Check EmploymentRetention
                alarming_str(excel,i,16)
                issues = issues +zero_one_accept_missing(excel,i,16)
                #Check JobTrainingEnrolled
                alarming_str(excel,i,17)
                issues = issues +zero_one_accept_missing(excel,i,17)
                #Check JobTrainingCompleted
                alarming_str(excel,i,18)
                issues = issues +zero_one_accept_missing(excel,i,18)
                #Check Savings, CreditScore, Debt
                issues = issues + check_int_float(excel,i,19)
                issues = issues + check_int_float(excel,i,20)
                issues = issues + check_int_float(excel,i,21)
                #Check AccessFinancialCoach
                alarming_str(excel,i,22)
                issues = issues +zero_one_accept_missing(excel,i,22)
                #Check Wages12 and Hours12 have the right value
                issues = issues + check_int_float(excel,i,23)
                issues = issues + check_int_float(excel,i,24)
                issues = issues + check_int_float(excel,i,25)
                issues = issues + check_int_float(excel,i,26)
                #Check IncomeClient, IncomeHH, PublicBenefits, ExpenseHH
                issues = issues + check_int_float(excel,i,27)
                issues = issues + check_int_float(excel,i,28)
                issues = issues + check_int_float(excel,i,29)
                issues = issues + check_int_float(excel,i,30)
                #Check BenefitsScreening: FinCap_Budget
                alarming_str(excel,i,31)
                issues = issues +zero_one_accept_missing(excel,i,31)
                alarming_str(excel,i,31)
                issues = issues +zero_one_accept_missing(excel,i,32)
                alarming_str(excel,i,33)
                issues = issues +zero_one_accept_missing(excel,i,33)
                alarming_str(excel,i,34)
                issues = issues +zero_one_accept_missing(excel,i,34)
                #Check FinCap_Goals
                alarming_str(excel,i,35)        
                if excel.cell(row = i, column = 35).value is not None:
                    a = mk_int(excel.cell(row = i, column = 35).value)
                    if a:
                        if a not in [0,1,2]:
                        #bold the cell
                            excel.cell(row = i, column = 35).fill = RED
                            excel.cell(row = i, column = 35).comment = number_out_of_range
                            excel.cell(row = i, column = last_column+issues+1).value = "{}:The number is out of range".format(excel.cell(row = 1, column = 35).value)
                            issues = issues + 1
                        else: pass
                    else: 
                        pass
                else:
                    pass
                #Check FinCap_EmergencyFund:FinCap_LateFee
                alarming_str(excel,i,36)
                issues = issues +zero_one_accept_missing(excel,i,36)
                alarming_str(excel,i,37)
                issues = issues +zero_one_accept_missing(excel,i,37)
                alarming_str(excel,i,38)
                issues = issues +zero_one_accept_missing(excel,i,38)
                alarming_str(excel,i,39)
                issues = issues +zero_one_accept_missing(excel,i,39)
                #Check Employment Status and Start Date
                issues = issues +employment_date_check(excel,i,14)
                issues = issues +employment_wage_check(excel,i,14)
                sheet_issues = sheet_issues + issues
                if issues !=0:
                    if total_discrepancy_column < issues:
                        total_discrepancy_column = issues
                    wrong_info +=1
                    for j in range(1, last_column+issues+1):
                        new_ws.cell(row = wrong_info+1,column = j).value = excel.cell(row = i,column = j).value
                        new_ws.cell(row = wrong_info+1,column = j).fill = copy(excel.cell(row = i,column = j).fill)
                        new_ws.cell(row = wrong_info+1,column = j).comment = copy(excel.cell(row = i,column = j).comment)
            if sheet_issues != 0:
                for number in range(1,total_discrepancy_column+1):
                    new_ws.cell(row = 1,column = last_column+number).value = "Discrepancy{}".format(number)
                for j in range(1, last_column+1):
                    new_ws.cell(row = 1,column = j).value = excel.cell(row = 1,column = j).value
                new_wb.save(f'{result_folder}/{Agency_Name}_Checked_Discrepancies.xlsx')
                wb.save(f'{result_folder}/{Agency_Name}_Checked_General.xlsx')
            DataLog.write("{} issues are detected in this data file.".format(sheet_issues)+"\n")
            DataLog.write("Columns that have data type issue: {}".format(data_type_issue)+"\n")
        else:
            DataLog.write("Not parse {} as the last column is not FinCap_LateFee.".format(Agency_Name)+"\n")
            print("Not parse "+Agency_Name)
        #document this file is not parsed
DataLog.close()


